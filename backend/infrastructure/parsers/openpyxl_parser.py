import openpyxl
from io import BytesIO
from typing import List
from .parse_result import ParseResult, EvaluationImportDTO, ValidationErreurDTO

class OpenpyxlParser:
    """Parseur robuste pour les fichiers de notes Excel."""

    def parser(self, contenu_fichier: bytes) -> ParseResult:
        resultat = ParseResult()
        try:
            wb = openpyxl.load_workbook(BytesIO(contenu_fichier), data_only=True)
            # On parcourt les onglets S5 et S6 si présents, sinon le premier
            onglets = [sheet for sheet in wb.sheetnames if sheet in ["S5", "S6"]]
            if not onglets:
                onglets = [wb.sheetnames[0]]

            for sheet_name in onglets:
                ws = wb[sheet_name]
                # On assume que la ligne 1 est l'en-tête
                # Format: Matricule | Nom | Prénom | Matière | CC | Examen | Rattrapage
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row): continue # Ligne vide

                    try:
                        dto = EvaluationImportDTO(
                            matricule=str(row[0]).strip(),
                            nom=str(row[1]).strip(),
                            prenom=str(row[2]).strip(),
                            matiere_libelle=str(row[3]).strip(),
                            note_cc=self._to_float(row[4]),
                            note_examen=self._to_float(row[5]),
                            note_rattrapage=self._to_float(row[6])
                        )
                        resultat.succes.append(dto)
                    except Exception as e:
                        resultat.erreurs.append(ValidationErreurDTO(
                            ligne=row_idx,
                            colonne="Toutes",
                            valeur=str(row),
                            message=str(e),
                            type_erreur="FORMAT"
                        ))
            return resultat
        except Exception as e:
            resultat.erreurs.append(ValidationErreurDTO(
                ligne=0,
                colonne="Fichier",
                message=f"Erreur fatale de lecture : {str(e)}",
                type_erreur="FATAL"
            ))
            return resultat

    def _to_float(self, value) -> Optional[float]:
        if value is None or str(value).strip() == "":
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
