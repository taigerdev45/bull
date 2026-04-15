import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO

class ExcelGenerator:
    """Générateur de classeurs Excel professionnels."""

    def generer_rapport_promotion(self, donnees: dict) -> bytes:
        wb = openpyxl.Workbook()
        
        # Onglet Synthèse
        ws_synth = wb.active
        ws_synth.title = "Synthèse"
        self._remplir_entete(ws_synth, ["Matricule", "Nom", "Prénom", "Moyenne Annuelle", "Décision", "Mention"])
        # Remplissage données synth...

        # Onglets S5 et S6
        for semestre in ["S5", "S6"]:
            ws = wb.create_sheet(title=semestre)
            self._remplir_entete(ws, ["Matricule", "Matière", "CC", "Examen", "Moyenne"])
            # Remplissage données...

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def _remplir_entete(self, ws, colonnes: List[str]):
        for col_idx, text in enumerate(colonnes, start=1):
            cell = ws.cell(row=1, column=col_idx, value=text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="092E20", end_color="092E20", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
